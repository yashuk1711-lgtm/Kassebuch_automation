import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

import config


def _to_float(value):
    """German-formatted amount string ("1.234,56") -> float, or None
    for blank/empty cells."""

    if value in (None, "") or (isinstance(value, float) and pd.isna(value)):
        return None

    return float(str(value).replace(".", "").replace(",", "."))


def write_xlsx(df, starting_balance, ending_balance, month, year, output_file):

    total_eingang = sum(v for v in (
        _to_float(x) for x in df["Eingang"]
    ) if v is not None)

    total_ausgang = sum(v for v in (
        _to_float(x) for x in df["Ausgang"]
    ) if v is not None)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month[:3]}{str(year)[-2:]}"

    ws["A1"] = f"Kassenbuch {config.RESTAURANT_NAME} {month}{year}"
    ws["A1"].font = Font(bold=True, size=14)

    ws["C4"] = "Anfangsbestand"
    ws["E4"] = round(starting_balance, 2)
    ws["F4"] = "Eingänge"
    ws["G4"] = round(total_eingang, 2)

    ws["C5"] = "Endebestand"
    ws["E5"] = round(ending_balance, 2)
    ws["F5"] = "Ausgänge"
    ws["G5"] = round(total_ausgang, 2)

    headers = [
        "Lfd.Nr", "Datum", "Beschreibung", "Zusatzlisches Beschreibung",
        "Eingang in €", "Ausgang in €", "Nuerbestand in €"
    ]

    header_row = 6

    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = Font(bold=True)

    row_num = header_row + 1

    for _, row in df.iterrows():

        ws.cell(row=row_num, column=1, value=row["Lfd.Nr"])

        date_cell = ws.cell(
            row=row_num,
            column=2,
            value=datetime.strptime(row["Datum"], "%d.%m.%Y")
        )
        date_cell.number_format = "DD.MM.YYYY"

        ws.cell(row=row_num, column=3, value=row["Beschreibung"])
        # column 4 (Zusatzlisches Beschreibung) intentionally left blank

        eingang = _to_float(row["Eingang"])
        if eingang is not None:
            ws.cell(row=row_num, column=5, value=eingang).number_format = "#,##0.00"

        ausgang = _to_float(row["Ausgang"])
        if ausgang is not None:
            ws.cell(row=row_num, column=6, value=ausgang).number_format = "#,##0.00"

        ws.cell(
            row=row_num, column=7, value=row["Neuestand"]
        ).number_format = "#,##0.00"

        row_num += 1

    widths = {"A": 8, "B": 12, "C": 24, "D": 24, "E": 14, "F": 14, "G": 16}

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(output_file)


def output_path(name):
    return os.path.join(config.OUTPUT_FOLDER, f"{config.MONTH_LOWER}_{name}.csv")


def safe_read_csv(path, columns, **kwargs):
    """Read a CSV, tolerating a missing file or one with zero rows
    (which pandas otherwise rejects as EmptyDataError since it has no
    header)."""

    if not os.path.exists(path):
        return pd.DataFrame(columns=columns)

    try:
        return pd.read_csv(path, **kwargs)
    except pd.errors.EmptyDataError:
        return pd.DataFrame(columns=columns)


def run():

    print("=== GENERATE_FULL_KASSENBUCH STARTED ===")

    if config.STARTING_BALANCE is None:
        raise SystemExit(
            "No starting balance found. Select a month (and, if this "
            "is the first month, enter a starting balance) before "
            "running the automation."
        )

    # Load files
    pos_df = safe_read_csv(
        output_path("pos"),
        ["date", "umsatz19", "umsatz7", "kartentrinkgeld"]
    )
    nexi_df = safe_read_csv(
        output_path("nexi"),
        ["date", "description", "amount"]
    )
    lieferando_cash_df = safe_read_csv(
        output_path("lieferando_cash"),
        ["date", "description", "amount"],
        sep=";"
    )
    lieferando_tips_df = safe_read_csv(
        output_path("lieferando_tips"),
        ["date", "description", "amount"],
        sep=";"
    )
    uber_df = safe_read_csv(
        output_path("uber"),
        ["date", "amount", "tip"],
        sep=";"
    )

    manual_expenses_file = os.path.join(
        config.MANUAL_FOLDER,
        f"{config.MONTH_LOWER}_manual_expenses.csv"
    )

    manual_df = safe_read_csv(
        manual_expenses_file,
        ["date", "description", "amount"],
        sep=";"
    )

    rows = []
    lfd_nr = 1
    starting_balance = config.STARTING_BALANCE
    kassenbuch_year = None

    for _, pos in pos_df.iterrows():

        date = str(pos["date"])

        if "/" in date:
            day, month, year = date.split("/")

        elif "-" in date:
            year, month, day = date.split("-")

        else:
            print("BAD DATE:", repr(date))
            continue

        formatted_date = f"{int(day)}.{int(month)}.{year}"
        day_month = f"{day.zfill(2)}.{month.zfill(2)}."
        kassenbuch_year = year

        # Umsatz 19%
        rows.append({
            "Lfd.Nr": lfd_nr,
            "Datum": formatted_date,
            "Beschreibung": "Umsatz 19%",
            "Eingang": pos["umsatz19"],
            "Ausgang": ""
        })
        lfd_nr += 1

        # Umsatz 7%
        rows.append({
            "Lfd.Nr": lfd_nr,
            "Datum": formatted_date,
            "Beschreibung": "Umsatz 7%",
            "Eingang": pos["umsatz7"],
            "Ausgang": ""
        })
        lfd_nr += 1

        # Karten Trinkgeld (always included, even when zero, to match
        # the original Kassenbuch format)
        rows.append({
            "Lfd.Nr": lfd_nr,
            "Datum": formatted_date,
            "Beschreibung": "Karten Trinkgeld",
            "Eingang": "",
            "Ausgang": pos["kartentrinkgeld"]
        })

        lfd_nr += 1

        # NEXI
        matching_nexi = nexi_df[
            nexi_df["date"] == day_month
        ]

        for _, nexi in matching_nexi.iterrows():

            rows.append({
                "Lfd.Nr": lfd_nr,
                "Datum": formatted_date,
                "Beschreibung": "NEXI GERMANY GMBH",
                "Eingang": "",
                "Ausgang": nexi["amount"]
            })

            lfd_nr += 1

        # Lieferando Cash
        matching_cash = lieferando_cash_df[
            lieferando_cash_df["date"] == day_month
        ]

        for _, cash in matching_cash.iterrows():

            rows.append({
                "Lfd.Nr": lfd_nr,
                "Datum": formatted_date,
                "Beschreibung": "Lieferando",
                "Eingang": cash["amount"],
                "Ausgang": ""
            })

            lfd_nr += 1

        # Lieferando Tips
        matching_tips = lieferando_tips_df[
            lieferando_tips_df["date"] == day_month
        ]

        for _, tip in matching_tips.iterrows():

            rows.append({
                "Lfd.Nr": lfd_nr,
                "Datum": formatted_date,
                "Beschreibung": "Lieferando Trinkgelder",
                "Eingang": "",
                "Ausgang": tip["amount"]
            })

            lfd_nr += 1

        # Uber Eats
        uber_date = f"{year}-{month}-{day}"

        matching_uber = uber_df[
            uber_df["date"] == uber_date
        ]

        for _, uber in matching_uber.iterrows():

            if pd.notna(uber.get("amount")) and str(uber.get("amount")) != "":

                rows.append({
                    "Lfd.Nr": lfd_nr,
                    "Datum": formatted_date,
                    "Beschreibung": "Uber Eats",
                    "Eingang": uber["amount"],
                    "Ausgang": ""
                })

                lfd_nr += 1

            if pd.notna(uber.get("tip")) and str(uber.get("tip")) != "":

                rows.append({
                    "Lfd.Nr": lfd_nr,
                    "Datum": formatted_date,
                    "Beschreibung": "Uber Eats Trinkgeld",
                    "Eingang": "",
                    "Ausgang": uber["tip"]
                })

                lfd_nr += 1

        # Manual Expenses
        matching_manual = manual_df[
            manual_df["date"] == day_month
        ]

        for _, expense in matching_manual.iterrows():

            rows.append({
                "Lfd.Nr": lfd_nr,
                "Datum": formatted_date,
                "Beschreibung": expense["description"],
                "Eingang": "",
                "Ausgang": expense["amount"]
            })

            lfd_nr += 1

    # Export
    df = pd.DataFrame(rows)

    balance = starting_balance
    balances = []

    for _, row in df.iterrows():

        income = row["Eingang"]
        expense = row["Ausgang"]

        if income != "":
            value = float(
                str(income)
                .replace(".", "")
                .replace(",", ".")
            )
            balance += value

        if expense != "":
            value = float(
                str(expense)
                .replace(".", "")
                .replace(",", ".")
            )
            balance -= value

        balances.append(round(balance, 2))

    df["Neuestand"] = balances

    os.makedirs(config.OUTPUT_FOLDER, exist_ok=True)

    output_file = os.path.join(
        config.OUTPUT_FOLDER,
        f"{config.MONTH_LOWER}_kassenbuch_full.csv"
    )

    df.to_csv(output_file, index=False)

    ending_balance = balances[-1] if balances else starting_balance

    xlsx_file = os.path.join(
        config.OUTPUT_FOLDER,
        f"{config.MONTH_LOWER}_kassenbuch_full.xlsx"
    )

    write_xlsx(
        df,
        starting_balance,
        ending_balance,
        config.MONTH,
        kassenbuch_year or datetime.now().year,
        xlsx_file
    )

    print("Created:", output_file)
    print("Created:", xlsx_file)
    print("Rows:", len(df))
    print("Anfangsbestand:", starting_balance)
    print("Endebestand:", ending_balance)


if __name__ == "__main__":
    run()
